import openpyxl
from mysql.connector import Error
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill, Protection
from openpyxl.utils import get_column_letter, column_index_from_string, rows_from_range
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl.worksheet.pagebreak import Break

import sport_categories
from db import connect_to_db, close_connect_db, update_db, insert_db, select_db
from sport_categories import sport_category, fina_points
from datetime import datetime

__distance_number__ = 1
__last_cell_of_page__ = 61
__last_break_in_distance_list__ = 49
__page_number__ = 1
swim_number = 1
row_for_break = 0
__index_row__ = 0


def list_swimmers_from_excel(file_name, table_name):
    count_swimmers = 0
    excel_file = openpyxl.load_workbook(file_name, data_only=True)
    sheet = excel_file.active

    # sheet_name = excel_file.sheetnames[0]
    # sheet = excel_file[sheet_name]
    conn = connect_to_db()

    query = f"SELECT * FROM tournaments WHERE table_name = '{table_name}'"

    data = select_db(conn, query)
    count_distances = []
    distances = []
    first_cell = 4
    for key, value in data.items():
        if key.startswith("distance"):
            if value is not None:
                distances.append(value)
                count_distances.append(first_cell + 5)
                first_cell += 5
    command = sheet['D10'].value
    if len(command) <= 1:
        close_connect_db(conn)
        return 'В файле заявки не указана команда. Исправьте файл и попробуйте заново.'
    query = f"SELECT second_name, first_name, gender, year, command,"
    for dis in range(0, len(distances)):
        query = query + ' distance_' + str(dis + 1) + ', distance_' \
                + str(dis + 1) + '_time, distance_' + str(dis + 1) + '_dsq,'
    query = query[:-1] + f" FROM {table_name} WHERE command = '{command}'"
    swimmers_in_db = select_db(conn, query, True)
    swimmers_in_zayavka = []
    text = f'*Заявка от команды {command}:*\n'
    text_changing = ''
    for i in range(15, 135):

        if sheet['E' + str(i)].value is None:
            continue
        swimmer = []
        cells = sheet['E' + str(i):'H' + str(i)]
        # print('cells', cells)
        for num, name in enumerate(cells[0]):
            # print(name)
            # print(name.value)
            name_1 = name.value
            if num == 0 or num == 1:
                # print('name_1', name_1)
                name_1 = name_1.strip()
                # print('name_1', name_1)
            if name_1 is not None:
                swimmer.append(name_1)
            else:
                close_connect_db(conn)
                return f"*ВНИМАНИЕ! Заявка не обработана.*\nНе полностью указаны данные участника под номером {sheet['D' + str(i)].value}", command
        swimmer.append(command)  # команда
        # print(swimmer)
        for number_column in count_distances:
            exh_column = get_column_letter(number_column)
            min_column = get_column_letter(number_column + 1)
            sec_column = get_column_letter(number_column + 2)
            ms_column = get_column_letter(number_column + 3)
            swimmer.append(sheet[exh_column + '12'].value)  # дистанция
            minute = 0
            sec = 0
            ms = 0
            if sheet[min_column + str(i)].value is not None:
                minute = int(sheet[min_column + str(i)].value)
            if sheet[sec_column + str(i)].value is not None:
                sec = int(sheet[sec_column + str(i)].value)
            if sheet[ms_column + str(i)].value is not None:
                ms = int(sheet[ms_column + str(i)].value)

            try:
                distance_time = minute * 60 + sec + ms / 100
                swimmer.append(distance_time)
            except:
                swimmer.append(None)
            if sheet[exh_column + str(i)].value == "л":
                swimmer.append('EXH')
            else:
                swimmer.append(None)
        swimmers_in_zayavka.append(tuple(swimmer))
    # print(swimmers_in_zayavka)
    swimmers_not_in_db = [x for x in swimmers_in_zayavka if x not in swimmers_in_db]
    swimmers_not_in_zayavka = [x for x in swimmers_in_db if x not in swimmers_in_zayavka]
    # print('swimmers_not_in_db', swimmers_not_in_db)
    # print('swimmers_not_in_zayavka', swimmers_not_in_zayavka)
    if len(swimmers_not_in_db) != 0:
        remove_from_swimmers_not_in_db = []
        for sw in swimmers_not_in_db:
            for swimmer_ in swimmers_in_db:
                if sw[0] == swimmer_[0] and sw[1] == swimmer_[1]:

                    query = f"INSERT INTO {table_name} (second_name, first_name, gender, year, command,"
                    for i in range(0, len(distances)):
                        query = query + ' distance_' + str(i + 1) + ', distance_' + str(i + 1) + '_time, distance_' \
                                + str(i + 1) + '_dsq,'
                    query = query[:-1] + ') VALUES ('
                    for h in sw:
                        if h == None:
                            query += 'NULL, '
                        elif h == "EXH":
                            query += "'EXH', "
                        elif isinstance(h, int) or isinstance(h, float):
                            query += str(h) + ", "
                        else:
                            query += "'" + str(h) + "', "
                    query = query[:-2] + f") ON DUPLICATE KEY UPDATE gender = '{sw[2]}', year = {sw[3]},"
                    for i in range(0, len(distances)):
                        query = query + ' distance_' + str(i + 1) + f"_time = {sw[(i + 1) * 3 + 3]}, distance_" + \
                                str(i + 1) + f'''_dsq = {'NULL' if sw[(i + 1) * 3 + 4]
                                                                   is None else "'" + sw[(i + 1) * 3 + 4] + "'"},'''
                    query = query[:-1]
                    try:
                        cursor = conn.cursor()
                        cursor.execute(query)
                        conn.commit()
                    except Error as error:
                        print(error)
                    finally:
                        cursor.close()
                    difference = []
                    difference.append(sw[0])
                    difference.append(sw[1])
                    difference.append(sw[3])
                    for num, x in enumerate(sw):
                        if x not in swimmer_:
                            for n in range(0, len(count_distances)):
                                if num == (n + 1) * 3 + 3:
                                    difference.append(sw[num - 1])
                                    difference.append(swimmer_[num])
                                    difference.append(x)
                    text_changing += '\n' + difference[0] + ' ' + difference[1] + ', ' + str(difference[2]) + ':'
                    for s in range(3, len(difference), 3):
                        text_changing += ' ' + difference[s] + ' - было ' + return_time(difference[s + 1]) + ' стало ' + \
                                         return_time(difference[s + 2]) + ','

                    text_changing = text_changing[:-1]
                    remove_from_swimmers_not_in_db.append(sw)
                    swimmers_not_in_zayavka.remove(swimmer_)
                    break
        new_swimmers_not_in_db = [x for x in swimmers_not_in_db if x not in remove_from_swimmers_not_in_db]
        if len(new_swimmers_not_in_db) != 0:
            text_add_swimmers = ''
            for swimmer__ in new_swimmers_not_in_db:
                query = f"INSERT INTO {table_name} (second_name, first_name, gender, year, command,"
                for i in range(0, len(distances)):
                    query = query + ' distance_' + str(i + 1) + ', distance_' + str(i + 1) + '_time, distance_' \
                            + str(i + 1) + '_dsq,'
                query = query[:-1] + ') VALUES ('
                for j in swimmer__:
                    if j == None:
                        query += 'NULL, '
                    elif isinstance(j, int) or isinstance(j, float):
                        query += str(j) + ", "
                    else:
                        query += "'" + str(j) + "', "

                query = query[:-2] + ')'
                try:
                    cursor = conn.cursor()
                    cursor.execute(query)
                    conn.commit()
                except Error as error:
                    if error.errno == 1062:
                        second_name = error.msg[error.msg.index("'") + 1:error.msg.index('-')]
                        first_name = error.msg[error.msg.index('-') + 1:error.msg.index('-', error.msg.index('-') + 1)]
                        name = second_name + ' ' + first_name
                        cursor.close()
                        query = f"DELETE FROM {table_name} WHERE second_name='{second_name}' " \
                                f"and first_name='{first_name}' and command='{command}'"
                        print(query)
                        try:
                            cursor = conn.cursor()
                            cursor.execute(query)
                            cursor.close()
                        except Error as er:
                            print(er)
                        text += f"*Добавлено участников: {str(count_swimmers)}*\n"
                        text += f"*ВНИМАНИЕ!*\n{str(count_swimmers + 1)}-й участник в заявке продублирован: " + name + \
                               ". Если это однофамильцы, то добавьте к участнику отчество в виде одной буквы. Например: 'О.'\n" \
                               "*И ЗАГРУЗИТЕ ЗАЯВКУ ЗАНОВО!*"
                        close_connect_db(conn)
                        return text, command

                finally:
                    cursor.close()
                text_add_swimmers += '\n' + swimmer__[0] + ' ' + swimmer__[1] + ', ' + str(swimmer__[3])
                count_swimmers += 1
            text += f"*Добавлено участников: {str(count_swimmers)}*\n"
            if count_swimmers <= 10:
                text += text_add_swimmers + '\n'

    if len(swimmers_not_in_zayavka) != 0:
        text += "*Убраны из турнира участники:*\n"
        for sw_ in swimmers_not_in_zayavka:
            query = f"DELETE FROM {table_name} WHERE second_name='{sw_[0]}' " \
                    f"and first_name='{sw_[1]}' and command='{sw_[4]}'"
            try:
                cursor = conn.cursor()
                cursor.execute(query)
                conn.commit()
            except Error as error:
                print(error)
            finally:
                cursor.close()
            text += sw_[0] + ' ' + sw_[1] + ', ' + str(sw_[3]) + '\n'
    if text_changing:
        text += '*Изменения в участниках:*' + text_changing
    if len(swimmers_not_in_zayavka) == 0 and len(swimmers_not_in_db) == 0:
        text += 'В заявке нет изменений'
    close_connect_db(conn)
    return text, command


def filter_swimmers(conn, ws, ws_track_list, year_1, year_2, gender, distance, last_swim, tracks, table_name,
                    organization=False):
    rows = []
    # conn = connect_to_db()
    distance_time = distance + '_time'
    distance_dsq = distance + '_dsq'
    query = f"SELECT second_name, first_name, gender, year, command, {distance_time}, {distance}, {distance_dsq} " \
            f"FROM {table_name} WHERE gender = '{gender}' AND (year BETWEEN {year_1} AND {year_2}) AND {distance_time} > 0 " \
            f"ORDER BY {distance_time}, RAND()"
    try:
        cursor = conn.cursor(buffered=True)
        cursor.execute(query)
        rows = cursor.fetchall()
    except:
        pass
    finally:
        cursor.close()
    # close_connect_db(conn)

    if rows:
        sorting_swimmers(ws, ws_track_list, rows, tracks, last_swim, year_1, year_2, organization)


def sorting_swimmers(ws, ws_track_list, rows, tracks, min_swimmers_in_swim, year_1, year_2, organization=False):
    global __distance_number__
    global __last_break_in_distance_list__
    global __page_number__
    quantity_swimmers = len(rows)
    quantity_swimmers_in_last_swim = len(rows) % tracks
    if quantity_swimmers_in_last_swim != 0:
        quantity_swims = len(rows) // tracks + 1
    else:
        quantity_swims = len(rows) // tracks
    last_but_one_swim = 0
    if quantity_swims == 1:
        last_swim = quantity_swimmers_in_last_swim
    elif quantity_swimmers_in_last_swim < min_swimmers_in_swim and quantity_swimmers_in_last_swim != 0:
        last_but_one_swim = tracks - (min_swimmers_in_swim - quantity_swimmers_in_last_swim)
        last_swim = min_swimmers_in_swim
    else:
        last_swim = quantity_swimmers_in_last_swim
    thins = Side(border_style="thin", )
    ws.append([''])
    ws.append([''])
    # print('rows', rows)
    for col in range(1, 8):
        ws.cell(row=ws.max_row, column=col).border = Border(top=thins)
    if rows[0][2] == 'ж':
        gender = 'девушки'
    else:
        gender = 'юноши'
    if year_1 == year_2:
        distance = rows[0][6] + ', ' + gender + ' ' + str(year_1) + ' г.р.'
    elif year_1 == '1950':
        distance = rows[0][6] + ', ' + gender + ' ' + str(year_2) + ' г.р. и старше'
    elif year_2 == '2025':
        distance = rows[0][6] + ', ' + gender + ' ' + str(year_1) + ' г.р. и старше'
    else:
        distance = rows[0][6] + ', ' + gender + ' ' + str(year_1) + '-' + str(year_2) + ' г.р.'

    ws['A' + str(ws.max_row + 1)] = 'Дистанция ' + str(__distance_number__)
    merging_cells = 'A' + str(ws.max_row) + ':C' + str(ws.max_row)
    ws.merge_cells(merging_cells)
    ws['D' + str(ws.max_row)] = distance
    merging_cells = 'D' + str(ws.max_row) + ':E' + str(ws.max_row)
    ws.merge_cells(merging_cells)
    ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='left')
    ws['D' + str(ws.max_row)].font = Font(name='Arial', size=11, bold=True)
    ws.append([''])
    for col in range(1, 8):
        ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
    for sheet in ws_track_list:
        if (rows[0][2] == 'ж' and sheet.max_row > 5) or (quantity_swims > 10 and sheet.max_row > 5):
            page_break = Break(id=sheet.max_row)
            sheet.row_breaks.append(page_break)
        sheet.append([''])
        sheet['A' + str(sheet.max_row + 1)] = 'Дистанция ' + str(__distance_number__)
        sheet.merge_cells('A' + str(sheet.max_row) + ':C' + str(sheet.max_row))
        sheet['D' + str(sheet.max_row)] = distance
        sheet.merge_cells('D' + str(sheet.max_row) + ':E' + str(sheet.max_row))
        sheet['D' + str(sheet.max_row)].alignment = Alignment(horizontal='left')
        sheet['D' + str(sheet.max_row)].font = Font(name='Arial', size=11, bold=True)
        sheet.append([''])
    __distance_number__ += 1

    for i in range(0, quantity_swims):
        track = []

        if i == quantity_swims - 2 and last_but_one_swim != 0:
            for k in range(i * tracks, i * tracks + last_but_one_swim):
                # print(k, rows[k])
                track.append(rows[k])
        elif i == quantity_swims - 1 and last_but_one_swim != 0:
            for k in range(quantity_swimmers - last_swim, quantity_swimmers):
                track.append(rows[k])
        else:
            for k in range(i * tracks, i * tracks + tracks):
                if k == quantity_swimmers:
                    break
                track.append(rows[k])
        # print('track', track)
        distance_swimmer = None
        if organization:
            distance_swimmer = track[0][6]
        if len(track) < tracks:
            for t in range(0, tracks - len(track)):
                track.append(())
        track = sorting_swims(track, tracks)
        for y in range(0, len(track)):
            try:
                # time_ms_str = str(int(track[y][5] * 100))
                # time_ms = time_ms_str[-2] + time_ms_str[-1]
                time_sec = return_time(track[y][5])
                # time_sec = str(int(track[y][5] // 60)) + "." + f'{int(track[y][5] % 60):02}' + "," + time_ms
                if track[y][7] is None:
                    track[y] = '', y + 1, track[y][0] + ' ' + track[y][1], track[y][3], track[y][4], time_sec,
                elif track[y][7] == 'EXH':
                    track[y] = '', y + 1, track[y][0] + ' ' + track[y][1] + ' (EXH)', track[y][3], track[y][
                        4], time_sec,
                else:
                    track[y] = '', y + 1, track[y][0] + ' ' + track[y][1], track[y][3], track[y][4], time_sec,
            except:
                track[y] = '', y + 1,
        # print('track222', track)
        write_excel(ws, ws_track_list, track, tracks, i + 1, quantity_swims, distance_swimmer, gender)
        __last_break_in_distance_list__ += quantity_swims * 2 + 3


def sorting_swims(track, quantity_tracks_in_swimming_pool):
    if quantity_tracks_in_swimming_pool == 4:
        track = [track[2], track[0], track[1], track[3]]
    elif quantity_tracks_in_swimming_pool == 5:
        track = [track[3], track[1], track[0], track[2], track[4]]
    elif quantity_tracks_in_swimming_pool == 6:
        track = [track[4], track[2], track[0], track[1], track[3], track[5]]
    elif quantity_tracks_in_swimming_pool == 7:
        track = [track[5], track[3], track[1], track[0], track[2], track[4], track[6]]
    elif quantity_tracks_in_swimming_pool == 8:
        track = [track[6], track[4], track[2], track[0], track[1], track[3], track[5], track[7]]
    return track


def write_excel(ws, ws_track_list, track, quantity_tracks_in_swimming_pool, swim_number_on_group, quantity_swims,
                distance=None, gender=None):
    global swim_number
    global __last_cell_of_page__
    thins = Side(border_style="thin", )
    ws.append([''])
    swim = '', '', 'Заплыв ' + str(swim_number_on_group) + ' из ' + str(
        quantity_swims),  # + ' (' + str(swim_number) + ')',
    max_row = ws.max_row
    list_of_cells_numbers = []
    for h in range(1, quantity_tracks_in_swimming_pool + 3):
        list_of_cells_numbers.append(h + max_row)
        # print('list_of_cells_numbers111', list_of_cells_numbers)
    if swim_number_on_group == 1:
        list_of_cells_numbers.append(list_of_cells_numbers[0] - 1)
        list_of_cells_numbers.append(list_of_cells_numbers[0] - 2)
        list_of_cells_numbers.append(list_of_cells_numbers[0] - 3)
        list_of_cells_numbers.append(list_of_cells_numbers[0] - 4)

    if __last_cell_of_page__ in list_of_cells_numbers:
        if swim_number_on_group == 1:
            page_break = Break(id=max_row - 5)
            __last_cell_of_page__ = max_row + 56

        else:
            page_break = Break(id=max_row)
            __last_cell_of_page__ = max_row + 61
        ws.row_breaks.append(page_break)

    elif list_of_cells_numbers[0] > __last_cell_of_page__:
        __last_cell_of_page__ = __last_cell_of_page__ + 61
    ws.append(swim)
    ws['C' + str(ws.max_row)].font = Font(name='Arial', underline="single", size=10)
    for sheet in ws_track_list:
        sheet.append(swim)
        sheet['C' + str(sheet.max_row)].font = Font(name='Arial', underline="single", size=10)
    swim_number += 1
    # print('swim_number', swim_number)
    for i in range(0, quantity_tracks_in_swimming_pool):
        ws.append(track[i])

        last_row = ws.max_row
        ws['D' + str(last_row)].alignment = Alignment(horizontal='center')
        ws['F' + str(last_row)].alignment = Alignment(horizontal='center')
        if distance:
            ws['N' + str(ws.max_row)] = distance
            ws['K' + str(ws.max_row)] = gender
        ws_track_list[i].append(track[i])
        ws_track_list[i]['F' + str(ws_track_list[i].max_row)] = '-'
        for col in range(2, 9):
            ws_track_list[i].cell(row=ws_track_list[i].max_row, column=col).border = Border(bottom=thins)
        ws_track_list[i]['D' + str(ws_track_list[i].max_row)].alignment = Alignment(horizontal='center')
        ws_track_list[i]['F' + str(ws_track_list[i].max_row)].alignment = Alignment(horizontal='center')
    # last_row = ws.max_row
    # thins = Side(border_style="thin", )
    # for row in range(last_row - quantity_tracks_in_swimming_pool + 1, last_row + 1):
    #     for col in range(1, 7):
    #         ws.cell(row=row, column=col).border = Border(top=thins, bottom=thins, left=thins, right=thins)


def create_start_protocol(last_swim, table_name):
    global swim_number
    global __last_cell_of_page__
    global __distance_number__

    now = datetime.now()
    sessions_list = {}
    wb = openpyxl.load_workbook('стартовый2.xlsx')
    ws = wb.active
    file_name_track_session_1 = ''
    file_name_track_session_2 = ''
    conn = connect_to_db()
    data = select_db(conn, None, False, 'temp_protocol', 'groups', 'gender', table_name="'" + table_name + "'")
    if '/' in data['groups']:
        sess_ = data['groups'].split('/')
        for num, value in enumerate(sess_, start=1):
            sessions_list[num] = value
    # print('sessions_list', sessions_list)
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    categories, sessions = tournament_year_categories(conn, table_name)

    ws['B2'] = data_tournament['name']
    ws['A5'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool']
    ws['F5'] = data_tournament['date'].strftime('%d.%m.%Y')
    sess_ = data['groups'].split('/')
    tracks = data_tournament['tracks']
    for num, value in enumerate(sess_, start=1):
        sessions_list[num] = value
    if sessions == 2:
        ws, ws2 = create_2_sheet(wb, ws, categories)

        wb_track_session_2 = openpyxl.load_workbook('дорожка.xlsx')
        ws_track_2 = wb_track_session_2.active
        ws_track_list_2 = [ws_track_2]
        for tr_ in range(1, tracks):
            ws_track_list_2.append(wb_track_session_2.copy_worksheet(ws_track_2))
            ws_track_list_2[tr_].title = str(tr_ + 1) + ' дорожка'
        for num, sheet in enumerate(ws_track_list_2, start=1):
            ws_centering(sheet)
            sheet.oddHeader.left.text = '&B' + str(num) + ' дорожка'
            sheet.oddHeader.center.text = data_tournament['name']
            sheet.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
            sheet.oddHeader.left.size = 13
            sheet.oddHeader.center.size = 9
            sheet.oddHeader.right.size = 9
            sheet.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M')  # + "  Страница &P из &N"
            sheet.oddFooter.right.size = 8
    wb_track = openpyxl.load_workbook('дорожка.xlsx')
    ws_track_1 = wb_track.active
    ws_track_list_1 = [ws_track_1]
    for tr in range(1, tracks):
        ws_track_list_1.append(wb_track.copy_worksheet(ws_track_1))
        ws_track_list_1[tr].title = str(tr + 1) + ' дорожка'
    for num, sheet in enumerate(ws_track_list_1, start=1):
        ws_centering(sheet)
        sheet.oddHeader.left.text = '&B' + str(num) + ' дорожка'  # + f"  Страница &P из &N "
        sheet.oddHeader.center.text = data_tournament['name']
        sheet.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
        sheet.oddHeader.left.size = 13
        sheet.oddHeader.center.size = 9
        sheet.oddHeader.right.size = 9
        sheet.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M')  # + "  Страница &P из &N"
        sheet.oddFooter.right.size = 8

    distances = []
    for key, value in data_tournament.items():
        if key.startswith("distance"):
            if value is not None:
                distances.append(value)
    if data['gender'] == 'ж':
        first_gender = 'ж'
        second_gender = 'м'
    else:
        first_gender = 'м'
        second_gender = 'ж'

    if data['groups'] is not None:
        for session, list_ in sessions_list.items():
            groups_list = list_.split('|')
            if session == 2:
                ws = ws2
                swim_number = 1
                ws_track_list = ws_track_list_2
                wb_track = wb_track_session_2
            else:
                ws_track_list = ws_track_list_1
            for i in range(1, len(distances) + 1):
                # print('IIIIIIIII', i)
                distance = 'distance_' + str(i)
                for list_of_years in groups_list:
                    years_list = list_of_years.split(',')

                    years_list.sort()
                    for _year_ in years_list:
                        # print('_year_22222', _year_)
                        if '+' in _year_:
                            # print('_year_', _year_)
                            min_year = '1950'
                            if len(years_list) == 1:
                                max_year = _year_[:-1]
                            else:
                                max_year = max(years_list)
                            # print('MAX YEAR', max_year)
                            break
                        elif '+' not in _year_ and '-' not in _year_:
                            min_year = min(years_list)
                            max_year = max(years_list)
                    for _year_ in years_list:
                        if '-' in _year_:
                            max_year = '2025'
                            if len(years_list) == 1:
                                min_year = _year_[:-1]
                            else:
                                min_year = years_list[0]
                            break
                        elif '+' not in years_list[0] and '-' not in years_list[0]:

                            min_year = min(years_list)
                            max_year = max(years_list)
                    # print(min_year, max_year)
                    # print(ws_track_list[0])
                    filter_swimmers(conn, ws, ws_track_list, min_year, max_year, first_gender, distance, int(last_swim),
                                    tracks, table_name)
                    filter_swimmers(conn, ws, ws_track_list, min_year, max_year, second_gender, distance, int(last_swim),
                                    tracks, table_name)
            __last_cell_of_page__ = 61

            ws_centering(ws)
            ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
            ws.oddFooter.right.size = 8
            ws.oddHeader.left.text = 'Стартовый протокол'
            ws.oddHeader.center.text = data_tournament['name']
            ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
            ws.oddHeader.left.size = 8
            ws.oddHeader.center.size = 8
            ws.oddHeader.right.size = 8
            if sessions == 2 and session == 1:
                file_name_track_session_1 = 'Дорожки_1_сессия_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
                wb_track.save(file_name_track_session_1)
            elif sessions == 2 and session == 2:
                file_name_track_session_2 = 'Дорожки_2_сессия_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
                wb_track.save(file_name_track_session_2)
            else:
                file_name_track = 'Дорожки_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
                wb_track.save(file_name_track)
    close_connect_db(conn)
    __distance_number__ = 1

    file_name = 'Стартовый_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    caption = '*Стартовый протокол ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    wb.save(file_name)
    swim_number = 1
    if sessions == 2:
        return file_name, caption, file_name_track_session_1, file_name_track_session_2
    else:
        return file_name, caption, file_name_track, file_name_track_session_2


def create_track_protocol_for_organization(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheets = wb.sheetnames

    thins = Side(border_style="thin", )
    dv = DataValidation(type="list", formula1='"DSQ,DNS,DNF,EXH"', allow_blank=True)
    formula = '"'
    for id_d in range(1, 56):
        formula += str(id_d) + ','
    formula = formula[:-1] + '"'
    dv_dsq = DataValidation(type="list", formula1=formula, allow_blank=True)

    for w in sheets:
        ws = wb[w]
        ws['A1'] = 'Стартовый протокол для организатора'
        ws.insert_cols(4, amount=1)
        if file_name.startswith("Стартовый"):
            rng = 6
        else:
            rng = 1
        for i in range(rng, ws.max_row + 1):
            if not isinstance(ws['E' + str(i)].value, int) and ws['E' + str(i)].value is not None:
                ws['C' + str(i + 1)] = 'DSQ - Дисквалификация, DNS - Неявка, DNF - Участник не финишировал, ' \
                                       'EXH - Лично. Вне конкурса'

                value_cell = ws['E' + str(i)].value
                distance = value_cell[:value_cell.index(',')]
                gender = value_cell[value_cell.index(',') + 2:value_cell.index(' ', value_cell.index(',') + 6)]
                ws['D' + str(i)] = ws['E' + str(i)].value
                ws['E' + str(i)] = ''
                ws.unmerge_cells('D' + str(i) + ':' + 'E' + str(i))
                ws['D' + str(i)].font = Font(name='Arial', size=11, bold=True)
                if gender == 'девушки':
                    gender = 'ж'
                else:
                    gender = 'м'
                ws['G' + str(i + 2)] = 'DSQ/DNS/DNF/EXH'
                ws['H' + str(i + 2)] = "мин"
                ws['I' + str(i + 2)] = 'сек'
                ws['J' + str(i + 2)] = "сотые"
                ws['L' + str(i + 2)] = "DSQ"
                ws['F' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['G' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['H' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['I' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['J' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['L' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['L' + str(i + 2)].alignment = Alignment(horizontal='center')
            elif ws['E' + str(i)].value is None and isinstance(ws['B' + str(i)].value, int):
                ws['C' + str(i)].protection = Protection(locked=False)
                ws['D' + str(i)].protection = Protection(locked=False)
                ws['E' + str(i)].protection = Protection(locked=False)
                ws['F' + str(i)].protection = Protection(locked=False)

            if ws['B' + str(i)].value:
                if ws['C' + str(i)].value is not None and 'EXH' in ws['C' + str(i)].value:
                    ws['G' + str(i)] = 'EXH'
                else:
                    ws['G' + str(i)] = ''
                ws['C' + str(i)].font = Font(bold=True)
                ws['O' + str(i)] = distance
                ws['D' + str(i)] = gender
                if file_name.startswith("Стартовый"):
                    ws['B' + str(i)].border = Border(bottom=thins)
                    ws['C' + str(i)].border = Border(bottom=thins)
                    ws['E' + str(i)].border = Border(bottom=thins)
                    ws['F' + str(i)].border = Border(bottom=thins)
                ws['D' + str(i)].border = Border(bottom=thins)
                ws['G' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
                ws['H' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
                ws['I' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
                ws['J' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
                ws['L' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
                ws['G' + str(i)] = '=IF(ISBLANK(L' + str(i) + '), "", "DSQ")'
                # ws['G' + str(i)].number_format = "_(* #,##0.00_);_(* (#,##0.00);_(* '-'??_);_(@_)"
                ws['G' + str(i)].protection = Protection(locked=False, hidden=True)
                ws['H' + str(i)].protection = Protection(locked=False)
                ws['I' + str(i)].protection = Protection(locked=False)
                ws['J' + str(i)].protection = Protection(locked=False)
                ws['L' + str(i)].protection = Protection(locked=False)
                # ws['J' + str(i)].fill = PatternFill('solid', fgColor="fce4d6")
                ws['M' + str(i)] = '=IF(AND(OR(ISBLANK(G' + str(i) + '),G' + str(i) + '=""),OR(ISBLANK(I' + str(i) + \
                                   '),ISBLANK(J' + str(i) + '))),"-",IF(ISBLANK(G' + str(i) + '),H' + str(i) + \
                                   '*60+I' + str(i) + '+J' + str(i) + '/100,G' + str(i) + '&H' + str(i) + \
                                   '*60+I' + str(i) + '+J' + str(i) + '/100))'
                ws['K' + str(i)] = '=IF(OR(ISBLANK(I' + str(i) + '),ISBLANK(J' + str(
                    i) + ')),"-",TEXT(H' + str(i) + ',"0")&"."&TEXT(I' + str(i) + ',"00")&","&TEXT(J' + str(
                    i) + ',"00"))'
                ws['K' + str(i)].protection = Protection(hidden=True)
                ws['G' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws['H' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws['I' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws['J' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws['L' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws['K' + str(i)].border = Border(bottom=thins)
                # ws['M' + str(i)].border = Border(left=double)
                ws['G' + str(i)].alignment = Alignment(horizontal='center')
                ws['H' + str(i)].alignment = Alignment(horizontal='center')
                ws['I' + str(i)].alignment = Alignment(horizontal='center')
                ws['J' + str(i)].alignment = Alignment(horizontal='center')
                ws['K' + str(i)].alignment = Alignment(horizontal='center')
                ws['L' + str(i)].alignment = Alignment(horizontal='center')
                # if val:
                #     ws['F' + str(i - 1)] = 'DSQ/DNS/DNF/EXH'
                #     ws['G' + str(i - 1)] = "мин"
                #     ws['H' + str(i - 1)] = 'сек'
                #     ws['I' + str(i - 1)] = "сот"
                #     ws['E' + str(i - 1)].font = Font(name='Arial', size=8)
                #     ws['F' + str(i - 1)].font = Font(name='Arial', size=8)
                #     ws['G' + str(i - 1)].font = Font(name='Arial', size=8)
                #     ws['H' + str(i - 1)].font = Font(name='Arial', size=8)
                #     ws['I' + str(i - 1)].font = Font(name='Arial', size=8)

                # val = False
                dv.add('G' + str(i))
                dv_dsq.add('L' + str(i))
            # else:
            #     val = True
        ws.column_dimensions['D'].width = 2
        ws.column_dimensions['E'].width = 5
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['H'].width = 5
        ws.column_dimensions['I'].width = 5
        ws.column_dimensions['J'].width = 5
        ws.column_dimensions['M'].hidden = True
        ws.column_dimensions['O'].hidden = True
        ws.column_dimensions['G'].width = 14
        ws.add_data_validation(dv)
        ws.add_data_validation(dv_dsq)
        ws.protection.password = 'IK135792468'
        ws.protection.enable()

    if file_name.startswith("Стартовый"):
        caption = "*Файл для записи результатов. В формате стартового протокола (по заплывам).*"
    else:
        caption = "*Файл для записи результатов. В формате файла для судей-секундометристов (по дорожкам).*"
    file_name = 'ОРГАНИЗАТОР_' + file_name
    wb.security.set_workbook_password("IK13579")
    wb.save(file_name)

    return file_name, caption

def create_start_protocol_for_organization(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheets = wb.sheetnames

    thins = Side(border_style="thin", )
    dv = DataValidation(type="list", formula1='"DSQ,DNS,DNF,EXH"', allow_blank=True)

    for w in sheets:
        ws = wb[w]
        ws['A1'] = 'Стартовый протокол для организатора'
        ws.insert_cols(4, amount=1)
        for i in range(6, ws.max_row + 1):
            if not isinstance(ws['E' + str(i)].value, int) and ws['E' + str(i)].value is not None:
                ws['C' + str(i + 2)] = 'DSQ - Дисквалификация, DNS - Неявка, DNF - Участник не финишировал, ' \
                                       'EXH - Лично. Вне конкурса'
                ws['C' + str(i + 2)].font = Font(name='Arial', size=10, bold=True)
                value_cell = ws['E' + str(i)].value
                distance = value_cell[:value_cell.index(',')]
                gender = value_cell[value_cell.index(',') + 3:value_cell.index(' ', value_cell.index(',') + 6)]
                ws['D' + str(i)] = ws['E' + str(i)].value
                ws['E' + str(i)] = ''
                ws.unmerge_cells('D' + str(i) + ':' + 'E' + str(i))
                ws['D' + str(i)].font = Font(name='Arial', size=11, bold=True)
                if gender == 'девушки':
                    gender = 'ж'
                else:
                    gender = 'м'
                ws['G' + str(i + 2)] = 'DSQ/DNS/DNF/EXH'
                ws['H' + str(i + 2)] = "мин"
                ws['I' + str(i + 2)] = 'сек'
                ws['J' + str(i + 2)] = "сот"
                ws['F' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['G' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['H' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['I' + str(i + 2)].font = Font(name='Arial', size=8)
                ws['G' + str(i + 2)].font = Font(name='Arial', size=8)
            elif ws['E' + str(i)].value is None and isinstance(ws['B' + str(i)].value, int):
                ws['C' + str(i)].protection = Protection(locked=False)
                ws['D' + str(i)].protection = Protection(locked=False)
                ws['E' + str(i)].protection = Protection(locked=False)
                ws['F' + str(i)].protection = Protection(locked=False)
        #
            if ws['B' + str(i)].value:
                if ws['C' + str(i)].value is not None and 'EXH' in ws['C' + str(i)].value:
                    ws['G' + str(i)] = 'EXH'
                else:
                    ws['G' + str(i)] = ''
        #         ws['C' + str(i)].font = Font(bold=True)
        #         ws['O' + str(i)] = distance
        #         ws['D' + str(i)] = gender
        #         ws['D' + str(i)].border = Border(bottom=thins)
        #         ws['G' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
        #         ws['H' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
        #         ws['I' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
        #         ws['J' + str(i)].fill = PatternFill('solid', fgColor="e2efda")
        #         ws['G' + str(i)].protection = Protection(locked=False)
        #         ws['H' + str(i)].protection = Protection(locked=False)
        #         ws['I' + str(i)].protection = Protection(locked=False)
        #         ws['J' + str(i)].protection = Protection(locked=False)
        #         # ws['J' + str(i)].fill = PatternFill('solid', fgColor="fce4d6")
        #         ws['M' + str(i)] = '=IF(AND(ISBLANK(G' + str(i) + '),OR(ISBLANK(I' + str(i) + \
        #                            '),ISBLANK(J' + str(i) + '))),"-",IF(ISBLANK(G' + str(i) + '),H' + str(i) + \
        #                            '*60+I' + str(i) + '+J' + str(i) + '/100,G' + str(i) + '&H' + str(i) + \
        #                            '*60+I' + str(i) + '+J' + str(i) + '/100))'
        #         ws['K' + str(i)] = '=IF(OR(ISBLANK(I' + str(i) + '),ISBLANK(J' + str(
        #             i) + ')),"-",TEXT(H' + str(i) + ',"0")&"."&TEXT(I' + str(i) + ',"00")&","&TEXT(J' + str(
        #             i) + ',"00"))'
        #
        #         ws['G' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        #         ws['H' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        #         ws['I' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        #         ws['J' + str(i)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
        #         ws['K' + str(i)].border = Border(bottom=thins)
        #         # ws['M' + str(i)].border = Border(left=double)
        #         ws['G' + str(i)].alignment = Alignment(horizontal='center')
        #         ws['H' + str(i)].alignment = Alignment(horizontal='center')
        #         ws['I' + str(i)].alignment = Alignment(horizontal='center')
        #         ws['J' + str(i)].alignment = Alignment(horizontal='center')
        #         ws['K' + str(i)].alignment = Alignment(horizontal='center')
        #         ws['L' + str(i)].alignment = Alignment(horizontal='left')
        #         # if val:
        #         #     ws['F' + str(i - 1)] = 'DSQ/DNS/DNF/EXH'
        #         #     ws['G' + str(i - 1)] = "мин"
        #         #     ws['H' + str(i - 1)] = 'сек'
        #         #     ws['I' + str(i - 1)] = "сот"
        #         #     ws['E' + str(i - 1)].font = Font(name='Arial', size=8)
        #         #     ws['F' + str(i - 1)].font = Font(name='Arial', size=8)
        #         #     ws['G' + str(i - 1)].font = Font(name='Arial', size=8)
        #         #     ws['H' + str(i - 1)].font = Font(name='Arial', size=8)
        #         #     ws['I' + str(i - 1)].font = Font(name='Arial', size=8)
        #
        #         # val = False
        #         dv.add('G' + str(i))
        #     # else:
        #     #     val = True
        # ws.column_dimensions['D'].width = 2
        # ws.column_dimensions['E'].width = 5
        # ws.column_dimensions['F'].width = 35
        # ws.column_dimensions['H'].width = 5
        # ws.column_dimensions['I'].width = 5
        # ws.column_dimensions['J'].width = 5
        # ws.column_dimensions['M'].hidden = True
        # ws.column_dimensions['O'].hidden = True
        # ws.column_dimensions['G'].width = 14
        # ws.add_data_validation(dv)
        # ws.protection.password = 'IK135792468'
        # ws.protection.enable()

    file_name = 'ОРГАНИЗАТОР_' + file_name
    caption = "*Файл для записи результатов. Дорожки.*"
    wb.save(file_name)

    return file_name, caption


def results_from_excel(file_name, table_name):
    # print('results fro excel')
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb.active
    distance = ''
    gender = ''
    second_name = ''
    first_name = ''
    command = ''
    result = 0.0
    year = 0
    conn = connect_to_db()
    for row in range(1, ws.max_row):
        value = str(ws['A' + str(row)].value)
        if 'ласты' in value:
            if 'девушки' in value:
                end_index = value.index('девушки')
                gender = 'ж'
            elif 'юноши' in value:
                end_index = value.index('юноши')
                gender = 'м'
            if value.startswith('200'):
                distance = '200м в/с кл.ласты'
            elif value.startswith('100'):
                distance = '100м в/с кл.ласты'
            elif value.startswith('50'):
                distance = '50м в/с кл.ласты'
            elif value.startswith('25'):
                distance = '25м в/с кл.ласты'
        elif value.startswith('200'):
            if 'девушки' in value:
                end_index = value.index('девушки')
                gender = 'ж'
            elif 'юноши' in value:
                end_index = value.index('юноши')
                gender = 'м'
            distance = '200м в/с'
        elif value.startswith('100'):
            if 'девушки' in value:
                end_index = value.index('девушки')
                gender = 'ж'
            elif 'юноши' in value:
                end_index = value.index('юноши')
                gender = 'м'
            distance = '100м в/с'
        value_b = ws['B' + str(row)].value
        # print(value_b)
        if value_b:
            value_b.strip()
            second_name = value_b[:value_b.index(' ')].strip()
            first_name = value_b[value_b.index(' ') + 1:].strip()
            command = ws['F' + str(row)].value
            result_str = ws['I' + str(row)].value
            year = int(ws['E' + str(row)].value)
            if result_str and ',' in result_str:
                min = int(result_str[:result_str.index('.')])
                sec = int(result_str[result_str.index('.') + 1:result_str.index(',')])
                ms = int(result_str[result_str.index(',') + 1:])
                result = min * 60 + sec + ms / 100
            else:
                result = 9999.0
        if value_b:
            # print(distance, gender, repr(second_name), repr(first_name), year, command, result)
            if distance == '200м в/с':
                set_query = 'distance_1'
            elif distance == '100м в/с':
                set_query = 'distance_2'
            elif distance == '200м в/с кл.ласты':
                set_query = 'distance_3'
            elif distance == '100м в/с кл.ласты':
                set_query = 'distance_4'
            elif distance == '50м в/с кл.ласты':
                set_query = 'distance_5'
            elif distance == '25м в/с кл.ласты':
                set_query = 'distance_6'
            query = f"INSERT INTO {table_name} (first_name, second_name, command, year, gender, " \
                    f"{set_query}, {set_query}_result, {set_query}_fina) " \
                    f"VALUES ('{first_name}', '{second_name}', '{command}', {year}, '{gender}', " \
                    f"'{distance}', {result}, {fina_points(gender, '25м', distance, result)}) " \
                    f"ON DUPLICATE KEY UPDATE {set_query}_result = {result}, {set_query} = '{distance}', " \
                    f"gender = '{gender}', {set_query}_fina = {fina_points(gender, '25м', distance, result)}"
            cursor = conn.cursor()
            cursor.execute(query)
            conn.commit()
    close_connect_db(conn)


def results_to_db(file_name, table_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    sheet_names = wb.sheetnames
    distances = []
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    swimming_pool = data_tournament['swimming_pool']
    for key, value in data_tournament.items():
        if key.startswith("distance"):
            if value is not None:
                distances.append(value)
    for sheet in sheet_names:
        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            result = ws['M' + str(row)].value

            if result is not None and result != '-' and result != 'EXH0' and ws['C' + str(row)].value is not None:
                # if result != '-' and result != 'EXH0':
                # print('result', result)
                distance = ws['O' + str(row)].value
                gender = ws['D' + str(row)].value
                name = ws['C' + str(row)].value
                second_name = name[:name.index(' ')]
                first_name = name[name.index(' ') + 1:]
                if "(EXH)" in first_name:
                    first_name = first_name[:first_name.index(" (EXH)")]
                command = ws['F' + str(row)].value

                number_distance = 0
                for i, value in enumerate(distances, start=1):
                    if value == distance:
                        number_distance = i
                if isinstance(result, float) or isinstance(result, int):
                    set_query = f"distance_{str(number_distance)}_result={result}, " \
                                f"distance_{str(number_distance)}_dsq=NULL, " \
                                f"distance_{str(number_distance)}_fina=" \
                                f"{fina_points(gender, swimming_pool, distance, result)}"
                else:
                    if result.startswith('EXH'):
                        result = float(result[3:].replace(',', '.'))

                        set_query = f"distance_{str(number_distance)}_result={result}, " \
                                    f"distance_{str(number_distance)}_dsq='EXH', " \
                                    f"distance_{str(number_distance)}_fina=" \
                                    f"{fina_points(gender, swimming_pool, distance, result)}"
                    else:
                        # result = 9999.0
                        set_query = f"distance_{str(number_distance)}_dsq='{result[:3]}', " \
                                    f"distance_{str(number_distance)}_result=9999, " \
                                    f"distance_{str(number_distance)}_fina=0"
                        if ws['L' + str(row)].value is not None:
                            set_query += f", distance_{str(number_distance)}_dsq_id={ws['L' + str(row)].value}"
                        print(set_query)
                query = f"UPDATE {table_name} SET {set_query} WHERE second_name='{second_name}' " \
                        f"AND first_name='{first_name}' AND command='{command}'"
                # print(query)
                update_db(conn, query)

                # query = f"SELECT id, gender, distance_{number_distance}_result FROM {table_name} " \
                #         f"WHERE first_name='{first_name}' and second_name='{second_name}' and command='{command}'"
                # data = select_db(conn, query, False)
                # print(data)
                # query = f"UPDATE {table_name} SET distance_{number_distance}_fina=" \
                #         f"{fina_points(data['gender'], swimming_pool, distance, result)} " \
                #         f"WHERE id={data['id']}"
                # print(query)
                # update_db(conn, query)

    # for number, distance in enumerate(distances, start=1):
    #     query = f"SELECT id, gender, distance_{number}_result, distance_{number}_dsq FROM {table_name} WHERE distance_{number}_result > 0 "
    #     data = select_db(conn, query, True)
    #     for swimmer in data:
    #         query = f"UPDATE {table_name} SET " \
    #                                 f"distance_{number}_fina={fina_points(swimmer[1], swimming_pool, distance, swimmer[2])} " \
    #                                 f"WHERE id={swimmer[0]}"
    #         update_db(conn, query)
    close_connect_db(conn)


def tournament_year_categories(conn, table_name):
    years = select_db(conn, None, False, 'tournaments', 'years', table_name="'" + table_name + "'")['years']
    sessions = 1
    session_categories = {}
    if '|' in years:
        sessions = 2
        session_categories[1] = years.split('|')[0]
        session_categories[2] = years.split('|')[1]
        session_categories[1] = session_categories[1].split(',')
        session_categories[2] = session_categories[2].split(',')
    else:
        session_categories[1] = years.split(',')
    for key, value in session_categories.items():
        categ = []
        for cat_ in value:
            categ.append(cat_.strip())
        session_categories[key] = categ
    return session_categories, sessions


def return_time(time_):
    time = "%.2f" % time_
    minute = str(int(time_ / 60))
    sec = f'{int(time_ % 60):02}'
    ms = str(time)[-2:]
    if minute == '0':
        return sec + '.' + ms
    else:
        return minute + ':' + sec + '.' + ms


def create_final_protocol(table_name, referi):
    global __index_row__
    wb = openpyxl.load_workbook("результаты.xlsx")
    ws = wb.active
    thins = Side(border_style="thin", )
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    categories, sessions = tournament_year_categories(conn, table_name)
    print(categories, sessions)
    for dist in range(1, data_tournament['quantity_dist'] + 1):
        distance = data_tournament[f"distance_{dist}"]
        for gender in ['ж', 'м']:
            if referi:
                for r in range(10, -1, -1):
                    if (ws.max_row + r) % 57 == 0:
                        for k in range(1, r):
                            ws.append([''])
                        break
            ws.append([''])
            distance_full_name = 'Дистанция ' + full_name_distance(distance) + ', ' + (
                'девушки' if gender == 'ж' else 'юноши')
            ws.append([distance_full_name])
            ws['A' + str(ws.max_row)].font = Font(bold=True, name='Arial', size=10)
            ws.merge_cells('A' + str(ws.max_row) + ':K' + str(ws.max_row))
            ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='left')
            for col in range(1, 12):
                ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
            if not distance.startswith("25"):
                text_normativ_1 = f"Норматив КМС: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['КМС'])} / " \
                                  f"Норматив I разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['I'])} / " \
                                  f"Норматив II разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['II'])} / " \
                                  f"Норматив III разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['III'])} "
                text_normativ_2 = f"Норматив I юн. разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['I(ю)'])} / " \
                                  f"Норматив II юн. разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['II(ю)'])} / " \
                                  f"Норматив III юн. разряд: {return_time(sport_categories.categories[gender][data_tournament['swimming_pool']][distance]['III(ю)'])} "
                ws.append([text_normativ_1])
            # ws.merge_cells('A' + str(ws.max_row) + ':K' + str(ws.max_row))
                ws['A' + str(ws.max_row)].font = Font(name='Arial', size=8)
                ws.append([text_normativ_2])
                ws['A' + str(ws.max_row)].font = Font(name='Arial', size=8)
                for col in range(1, 12):
                    ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
            for sess in range(1, sessions + 1):
                for cat_ in categories[sess]:
                    # print(cat_)
                    no_cat = False
                    year_1, year_2, name_cat = years_in_categories(cat_)

                    for r in range(6, -1, -1):
                        if (ws.max_row + r) % 57 == 0:

                            for k in range(0, r):
                                ws.append([''])
                            ws.append([distance_full_name])
                            for col in range(1, 12):
                                ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
                            break

                    query = f"SELECT second_name, first_name, gender, year, command, distance_{dist}_result, " \
                            f"distance_{dist}_dsq, distance_{dist}_fina, distance_{dist} " \
                            f"FROM {table_name} WHERE (year BETWEEN {year_1} AND {year_2}) AND gender = '{gender}' " \
                            f"AND distance_{dist}_result > 0 ORDER BY distance_{dist}_result"
                    # print(query)
                    data = select_db(conn, query, True)
                    # print(data)
                    if data:
                        ws.append([''])
                        ws.append(['Год рождения ' + name_cat])
                        new_data = []
                        for swimmer in data:
                            if swimmer[6] == 'EXH':
                                new_data.append(swimmer)
                        if new_data:
                            for swimmer in new_data:
                                if swimmer in data:
                                    data.remove(swimmer)
                                    data.append(swimmer)
                        create_pre_excel(ws, data, gender, name_cat, data[0][8], data_tournament, True, referi, no_cat,
                                         distance_full_name)




    sheet_headers(ws, data_tournament, 'Итоговый протокол', 'Итоговый протокол')

    caption = '*Итоговый протокол*\n' + \
              data_tournament['date'].strftime('%d.%m.%Y') + ', ' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + data_tournament['swimming_pool']

    file_name = 'Итоговый_протокол_' + data_tournament['date'].strftime(
        '%d_%m_%Y') + '.xlsx'

    if referi:
        refiries = select_db(conn, None, False, 'temp_protocol', 'groups', table_name="'" + table_name + "'")['groups'].split('|')
        query = f"DELETE FROM `temp_protocol` WHERE table_name = '{table_name}'"
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        for ind in range(57, ws.max_row, 57):
            add_footer_to_final_protocol(ws, ind, data_tournament, thins, refiries)
        if ws.max_row % 57 == 0 or (ws.max_row + 1) % 57 or (ws.max_row + 2) % 57:
            pass
        else:
            while ws.max_row % 57 == 0:
                ws.append([''])
            add_footer_to_final_protocol(ws, ws.max_row, data_tournament, thins, refiries)
    close_connect_db(conn)
    wb.save(file_name)

    return file_name, caption


def add_footer_to_final_protocol(ws, ind, data_tournament, thins, refiries):
    for col in range(1, 12):
        ws.cell(row=ind - 2, column=col).border = Border(bottom=thins)
    ws['A' + str(ind - 1)] = 'Бассейн ' + data_tournament['swimming_pool'] + ', ' + \
                             str(data_tournament['tracks']) + ' дор.'
    ws['F' + str(ind - 1)] = data_tournament['place']
    ws['K' + str(ind - 1)] = data_tournament['date'].strftime('%d.%m.%Y')
    ws['A' + str(ind - 1)].font = Font(name='Arial', size=9)
    ws['F' + str(ind - 1)].font = Font(name='Arial', size=9)
    ws['K' + str(ind - 1)].font = Font(name='Arial', size=9)
    ws['F' + str(ind - 1)].alignment = Alignment(horizontal='center')
    ws['K' + str(ind - 1)].alignment = Alignment(horizontal='right')

    ws['A' + str(ind)] = 'Гл.судья: ' + refiries[0]
    ws['F' + str(ind)] = 'Рефери: ' + refiries[1]
    ws['K' + str(ind)] = 'Гл.секретарь: ' + refiries[2]

    ws['A' + str(ind)].font = Font(name='Arial', size=9)
    ws['F' + str(ind)].font = Font(name='Arial', size=9)
    ws['K' + str(ind)].font = Font(name='Arial', size=9)
    ws['F' + str(ind)].alignment = Alignment(horizontal='center')
    ws['K' + str(ind)].alignment = Alignment(horizontal='right')
    for col in range(1, 12):
        ws.cell(row=ind, column=col).border = Border(bottom=thins)


def create_pre_results(table_name, distance, gender, session):
    wb = openpyxl.load_workbook("результаты.xlsx")
    ws = wb.active
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    categories, sessions = tournament_year_categories(conn, table_name)
    not_categories = \
        select_db(conn, None, False, 'temp_protocol', 'groups', table_name="'" + table_name + "'")
    no_categ = {}
    if not_categories is not None:
        all_not_cat = not_categories['groups'].split('|')
        for c in all_not_cat:
            cc = c.split('_')
            no_categ[cc[0]] = cc[1], cc[2]
        # print(no_categ)
    query = f"DELETE FROM `temp_protocol` WHERE table_name = '{table_name}'"
    cursor = conn.cursor()
    cursor.execute(query)
    conn.commit()

    for key, value in data_tournament.items():
        if value == distance:
            dist_column_name = key
            break
    years = []
    for cat_ in categories[int(session)]:
        no_cat = False
        year_1, year_2, name_cat = years_in_categories(cat_)
        years.append(int(year_1))
        years.append(int(year_2))
        query = f"SELECT second_name, first_name, gender, year, command, {dist_column_name}_result, " \
                f"{dist_column_name}_dsq, {dist_column_name}_fina, id " \
                f"FROM {table_name} WHERE (year BETWEEN {year_1} AND {year_2}) AND gender = '{gender}' " \
                f"AND {dist_column_name}_result > 0 ORDER BY {dist_column_name}_result"
        data = select_db(conn, query, True)
        if data:
            new_data = []
            for swimmer in data:
                if swimmer[6] == 'EXH':
                    new_data.append(swimmer)
            if new_data:
                for swimmer in new_data:
                    if swimmer in data:
                        data.remove(swimmer)
                        data.append(swimmer)

            if not_categories and cat_ in no_categ.keys():
                # print(no_categ[cat_])
                if no_categ[cat_][0] == distance and no_categ[cat_][1] == gender:
                    no_cat = True
            create_pre_excel(ws, data, gender, name_cat, distance, dist_column_name, data_tournament, False, False, no_cat, None)
    close_connect_db(conn)
    age = ''
    if years:
        if min(years) == 1950:
            age = str(max(years)) + ' и старше'
        elif max(years) == 2025:
            age = str(min(years)) + ' и младше'
        else:
            age = str(min(years)) + '-' + str(max(years))
    sheet_headers(ws, data_tournament, 'Результаты',
                  'Результаты ' + distance + (', девушки, ' if gender == 'ж' else ', юноши, ') + age)
    caption = '*Результаты ' + distance + (', девушки, ' if gender == 'ж' else ', юноши, ') + age + '*\n\n' + \
              data_tournament['date'].strftime('%d.%m.%Y') + ', ' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + data_tournament['swimming_pool']

    dist_ = distance.replace('/', '').replace(' ', '_')

    if age:
        age = age.replace(' ', '_')
    file_name = 'Результаты_' + dist_ + ('_девушки_' if gender == 'ж' else '_юноши_') + age + data_tournament[
        'date'].strftime('_%d_%m_%Y') + '.xlsx'
    wb.save(file_name)

    return file_name, caption


def create_medals(table_name, file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    conn = connect_to_db()
    categories, sessions = tournament_year_categories(conn, table_name)
    categories_ = []
    for cat_ in categories.values():
        for c in cat_:
            categories_.append(c)
    distances = []
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    for key, value in data_tournament.items():
        if key.startswith("distance") and value is not None:
            distances.append(value)
    sheet_headers(ws, data_tournament, 'Медалисты', 'Медалисты')
    query = f"UPDATE {table_name} SET `gold`=0,`silver`=0,`bronze`=0 WHERE 1"
    update_db(conn, query)
    for i, distance in enumerate(distances, start=1):
        for cat_ in categories_:
            year_1, year_2, name_cat = years_in_categories(cat_)
            for gender in ['ж', 'м']:
                query = f"SELECT `id`, `distance_{i}_result` FROM {table_name} WHERE (distance_{i}_dsq = '' " \
                        f"OR distance_{i}_dsq IS NULL) AND (year BETWEEN {year_1} AND {year_2}) " \
                        f"AND gender = '{gender}' " \
                        f"AND distance_{i}_result IS NOT NULL ORDER BY distance_{i}_result LIMIT 5"
                data = select_db(conn, query, True)
                if data:
                    multimedals(conn, table_name, data)

    medals_excel(conn, ws, table_name)
    close_connect_db(conn)
    caption = '*Медалисты ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'

    file_name = 'Медалисты_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    wb.save(file_name)

    return file_name, caption


def ws_centering(ws):
    ws.page_margins = PageMargins(left=0.39, right=0.39, top=0.55, bottom=0.55, header=0.31, footer=0.31)
    ws.print_options = PrintOptions(horizontalCentered=True)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False


def multimedals(conn, table_name, data):
    gold = []
    silver = []
    bronze = []
    numbers = []
    times = []
    for dat in data:
        numbers.append(dat[0])
        times.append(dat[1])

    if data:
        index = 0
        idx = [x[0] for x in enumerate(times) if x[1] == times[0]]

        index += len(idx)
        for n in idx:
            gold.append(numbers[n])
        try:
            idx = [x[0] for x in enumerate(times) if x[1] == times[index]]
            for n in idx:
                silver.append(numbers[n])
        except:
            pass
        index += len(idx)
        try:
            idx = [x[0] for x in enumerate(times) if x[1] == times[index]]
            for n in idx:
                bronze.append(numbers[n])
        except:
            pass

        for g in gold:
            query = f"UPDATE {table_name} SET gold = gold + 1 WHERE id = {g}"
            update_db(conn, query)
        for s in silver:
            query = f"UPDATE {table_name} SET silver = silver + 1 WHERE id = {s}"
            update_db(conn, query)
        for b in bronze:
            query = f"UPDATE {table_name} SET bronze = bronze + 1 WHERE id = {b}"
            update_db(conn, query)


def medals_excel(conn, ws, table_name):
    query = f"SELECT second_name, first_name, year, command, gold, silver, bronze FROM {table_name} " \
            f"WHERE gold != 0 OR silver != 0 OR bronze != 0 ORDER BY gold DESC, silver DESC, bronze DESC, second_name"
    data = select_db(conn, query, True)
    # print(data)

    for number, swimmer in enumerate(data, start=1):
        swimmer_ = []
        swimmer_.append(number)
        swimmer_.append(swimmer[0] + ' ' + swimmer[1])
        swimmer_.append(swimmer[2])
        swimmer_.append(swimmer[3])
        swimmer_.append(swimmer[4])
        swimmer_.append(swimmer[5])
        swimmer_.append(swimmer[6])
        swimmer_.append(swimmer[4] + swimmer[5] + swimmer[6])
        try:
            if swimmer_[4] == data[number - 2][4] and swimmer_[5] == data[number - 2][5] \
                    and swimmer_[6] == data[number - 2][6]:
                swimmer_[0] = ''
        except:
            pass
        ws.append(swimmer_)
        ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['E' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['G' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['H' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        if isinstance(ws['A' + str(ws.max_row)].value, int):
            for col in range(1, 9):
                ws.cell(row=ws.max_row, column=col).border = Border(top=Side(border_style="thin", ))


def create_pre_excel(ws, data, gender, name_cat, distance, dist_column_name, data_tournament, final, referi, no_cat,
                     distance_full_name=None):
    distance_for_category = distance

    thins = Side(border_style="thin", )

    if not final:
        ws.append([''])
        for col in range(1, 12):
            ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
        ws.append([distance + ', ' + ('девушки' if gender == 'ж' else 'юноши') + ', ' + name_cat])
        ws.merge_cells('A' + str(ws.max_row) + ':K' + str(ws.max_row))
        ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        for col in range(1, 12):
            ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
        ws.append([''])

    if not no_cat:
        ws.append(['', 'Место', 'Фамилия Имя', '', 'г.р.', 'Команда', '', 'Итоговый', 'Разряд', 'FINA2022'])
    else:
        ws.append(['', '', 'Фамилия Имя', '', 'г.р.', 'Команда', '', 'Итоговый', 'Разряд', 'FINA2022'])
    for col in range(1, 12):
        ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8)
        ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical='center')
        if col == 5 or col == 8:
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center',
                                                                      vertical='center')
    add_swimmer_to_excel(ws, data, data_tournament, distance_for_category, dist_column_name, final, referi, no_cat, distance_full_name)


def add_swimmer_to_excel(ws, data, data_tournament, distance_for_category, dist_column_name, final, referi, no_cat,
                         distance_full_name=None):
    thins = Side(border_style="thin", )
    for number, swimmer_ in enumerate(data, start=1):
        print(swimmer_)
        swimmer = ['']
        # last_row = ws.max_row
        # last_place = ws['B' + str(last_row)].value
        # print('last_row', last_row)
        # print('last_place', last_place)
        if not no_cat:
            # if last_place is not None and last_place != '' and isinstance(last_place, int):
            if swimmer_[5] == data[number - 2][5] and len(data) > 1:
                swimmer.append('')
                # elif last_place != '':
                #     swimmer.append(int(last_place) + 1) #int(last_place) + 1
            else:
                swimmer.append(number)
        else:
            swimmer.append('')

        swimmer.append(swimmer_[0] + ' ' + swimmer_[1])
        swimmer.append('')
        swimmer.append(swimmer_[3])
        swimmer.append(swimmer_[4])
        swimmer.append('')
        if swimmer_[5] and swimmer_[5] != 9999.0 or swimmer_[6] == 'EXH':
            # time = "%.2f" % swimmer_[5]
            # min = str(int(swimmer_[5] / 60))
            # sec = f'{int(swimmer_[5] % 60):02}'
            # ms = str(time)[-2:]
            if swimmer_[6] == 'EXH':
                swimmer[1] = 'EXH'
            swimmer.append(return_time(swimmer_[5]))
        elif swimmer_[6]:
            swimmer[1] = swimmer_[6]
            swimmer.append('')
        else:
            swimmer.append('-')
        swimmer.append(
            sport_category(swimmer_[2], data_tournament['swimming_pool'], distance_for_category, swimmer_[5]))
        swimmer.append(swimmer_[7])
        if final:
            if referi and (ws.max_row + 3) % 57 == 0:
                ws.append([''])
                ws.append([''])
                ws.append([''])
                ws.append([distance_full_name])
                for col in range(1, 12):
                    ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
            elif not referi and ws.max_row % 57 == 0:
                ws.append([distance_full_name])
                for col in range(1, 12):
                    ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
        ws.append(swimmer)
        if swimmer[1] == 'DSQ':
            conn = connect_to_db()
            query = f"SELECT b.name FROM {data_tournament['table_name']} a, dsq b, {data_tournament['table_name']} c " \
                    f"WHERE c.id = {swimmer_[8]} AND a.{dist_column_name}_dsq_id=b.id"
            data = select_db(conn, query)
            close_connect_db(conn)
            print(data)
            if data is not None:
                print(len(data['name']))
                ws.append(['', '', data['name']])

                ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='left', wrapText=True)
                ws['C' + str(ws.max_row)].font = Font(name='Arial', size=9, italic=True)
                ws.merge_cells('C' + str(ws.max_row) + ':J' + str(ws.max_row))
                if len(data['name']) > 100:
                    ws.row_dimensions[ws.max_row].height = 25.5
        else:
            last_row = ws.max_row
            # thins = Side(border_style="thin", )
            # for col in range(1, 8):
            #     ws.cell(row=last_row, column=col).border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws['B' + str(last_row)].alignment = Alignment(horizontal='center')
            ws['E' + str(last_row)].alignment = Alignment(horizontal='center')
            ws['H' + str(last_row)].alignment = Alignment(horizontal='center')
            ws['I' + str(last_row)].alignment = Alignment(horizontal='center')
            ws['J' + str(last_row)].alignment = Alignment(horizontal='center')
        # if not final and number <= 6:
        #     # print('NUMBER', number)
        #     if swimmer[1] == 1:
        #         color = 'ffe8a7'
        #     elif swimmer[1] == 2:
        #         color = 'e4e4e4'
        #     elif swimmer[1] == 3:
        #         color = 'efd7bf'
        #     else:
        #         color = 'ffffff'
        #
        #     ws['B' + str(last_row)].fill = PatternFill('solid', fgColor=color)
        # ws['C' + str(last_row)].fill = PatternFill('solid', fgColor=color)
        # ws['D' + str(last_row)].fill = PatternFill('solid', fgColor=color)
        # ws['E' + str(last_row)].fill = PatternFill('solid', fgColor=color)
        # ws['F' + str(last_row)].fill = PatternFill('solid', fgColor=color)
        # ws['G' + str(last_row)].fill = PatternFill('solid', fgColor=color)


def create_app(table_name):
    file = 'заявка.xlsx'
    excel_file = openpyxl.load_workbook(file)
    sheet_name = excel_file.sheetnames[0]
    sheet = excel_file[sheet_name]
    conn = connect_to_db()
    data = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    # print(data)
    sheet['F2'] = data['name']
    sheet['F3'] = data['date'].strftime('%d.%m.%Y')
    sheet['F4'] = data['place'] + ', ' + data['swimming_pool'] + ', ' + str(data['tracks']) + ' дор.'
    column_index = 4
    distance_count = 0
    for i in range(1, 11):
        if data['distance_' + str(i)] is not None:
            cell = get_column_letter(column_index + 5) + '12'
            sheet[cell] = data['distance_' + str(i)]
            # sheet[cell].font = Font(bold=True)
            column_index += 5
            distance_count += 1
        else:
            sheet.delete_cols(column_index + 5, 5)
    # if distance_count == 1:
    #     text_for_app = 'Для заполнения доступны только зеленые поля.\nПосле ввода данных в ячейки мин, сек, сотые - ' \
    #                    'в ячейке "Результат введеного времени" автоматически появится заявленное время. ' \
    #                    'Проверьте его, правильно ли отобразилось время участника в формате: мин.сек,сотые\n' \
    #                    'Если участник НЕ плывет какую-либо дистанцию - оставьте ячейки "Время" на данной дистанции пустыми.\n' \
    #                    'Если участник плывет дистанцию лично (вне конкурса) - поставьте в столбец EXH букву "л".\n' \
    #                    'Если участник снимается с соревнований ДО формирования стартового протокола, ' \
    #                    'удалите все данные спортсмена и отправьте обновленную заявку организатору.\n' \
    #                    'Если есть 2 и более спортсмена с одинаковыми фамилиями и именами, в ячейке ' \
    #                    'Имя вместе с именем укажите первую букву отчества (О.)\nОграничения на вводимые данные:\n' \
    #                    'Название команды - до 40 символов\nТип населенного пункта - выберите из списка; ' \
    #                    'Название населеленного пункта - до 20 символов\nФамилия и Имя - до 30 символов; ' \
    #                    'Пол м/ж - только буква "м" или "ж"; г.р. - число от 1950 до 2022\nEXH - буква "л" или пустое; ' \
    #                    'мин,сек - число от 0 до 59 или пустое; сотые - число от 0 до 99 или пустое ' \
    #                    '(пустое поле будет означать 0).'
    #     sheet['N1'] = text_for_app
    #     sheet['N1'].alignment = Alignment(wrapText=True, vertical='top')
    #     sheet['N1'].font = Font(size=9)
    sheet.protection.password = 'IK135792468'
    sheet.protection.enable()
    file_name = 'Заявка_' + data['date'].strftime('%d_%m_%Y') + '.xlsx'
    excel_file.save(file_name)
    caption = '*Техническая заявка ' + data['date'].strftime('%d.%m.%Y') + '*\n' + \
              data['name'] + '\n' + data['place'] + ', ' + \
              data['swimming_pool'] + ', ' + str(data['tracks']) + ' дор.\n'
    return file_name, caption


def create_2_sheet(wb, ws, categories):
    ws2 = wb.copy_worksheet(ws)
    categ1 = []
    categ2 = []
    plus = False
    minus = False
    for cat_ in categories[1]:
        if '-' in cat_:
            categ1.append(cat_[:4])
            categ1.append(cat_[5:])
        else:
            categ1.append(cat_)
        if '+' in cat_:
            plus = True
    # print(categ1)
    min_year = min(categ1)
    max_year = max(categ1)
    if plus:
        ws['A4'] = '1 сессия, ' + max_year + ' г.р. и старше'
    else:
        ws['A4'] = '1 сессия, ' + min_year + '-' + max_year + ' г.р.'

    for cat_ in categories[2]:
        if '-' in cat_ and len(cat_) > 6:
            categ2.append(cat_[:4])
            categ2.append(cat_[5:])
        else:
            categ2.append(cat_)
        if '-' in cat_ and len(cat_) < 6:
            minus = True
    # print(categ2)
    min_year = min(categ2)
    max_year = max(categ2)
    if minus:
        ws2['A4'] = '2 сессия, ' + min_year + ' г.р. и младше'
    else:
        ws2['A4'] = '2 сессия, ' + min_year + '-' + max_year + ' г.р.'

    ws['A4'].font = Font(name='Arial', size=10, bold=True)
    ws2['A4'].font = Font(name='Arial', size=10, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    ws2['A4'].alignment = Alignment(horizontal='center')
    ws.title = "1 сессия"
    ws2.title = "2 сессия"
    return ws, ws2


def sheet_headers(ws, data_tournament, name_header, name):
    now = datetime.now()
    ws.oddHeader.center.text = data_tournament['name']
    ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
    ws.oddHeader.left.text = name_header
    ws.oddHeader.center.size = 8
    ws.oddHeader.right.size = 8
    ws.oddHeader.left.size = 8
    ws['A1'] = name
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., ' + data_tournament['date'].strftime('%d.%m.%Y')

    ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
    ws.oddFooter.right.size = 8
    ws_centering(ws)


def full_name_distance(distance):
    dist = ''
    if 'ласты' in distance:
        dist = 'вольный стиль классические ласты'
    elif 'в/с' in distance:
        dist = 'вольный стиль'
    elif 'спин' in distance or 'брасс' in distance:
        return distance
    elif 'батт' in distance:
        dist = 'баттерфляй'
    elif 'комплекс' in distance:
        dist = 'комплексное плавание'
    return distance[:distance.index(' ') + 1] + dist


def create_list_2(table_name):
    file = 'заявочный.xlsx'
    now = datetime.now()
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    ws.oddHeader.center.text = data_tournament['name']
    ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
    ws.oddHeader.left.text = "Заявочный протокол"
    ws.oddHeader.center.size = 8
    ws.oddHeader.right.size = 8
    ws.oddHeader.left.size = 8
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., ' + data_tournament['date'].strftime('%d.%m.%Y')
    distances = []
    thins = Side(border_style="thin", )
    for key, value in data_tournament.items():
        if key.startswith("distance") and value is not None:
            distances.append(value)
    for i, distance in enumerate(distances, start=1):
        distance_full_name = full_name_distance(distance)
        for gender in ['ж', 'м']:
            query = f"SELECT GROUP_CONCAT(second_name, ' ', first_name) AS name, gender, year, " \
                    f"command, distance_{i}_time, distance_{i}_dsq " \
                    f"FROM {table_name} WHERE (year BETWEEN 1940 AND 2022) AND gender = '{gender}' " \
                    f"AND distance_{i}_time > 0 GROUP BY second_name, first_name, year, command, " \
                    f"distance_{i}_time, distance_{i}_dsq ORDER BY distance_{i}_time"
            data = select_db(conn, query, True)
            # print(data)
            if data:
                ws.append([''])
                ws.append([''])
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).border = Border(top=thins)
                ws.append([('Девушки' if gender == 'ж' else 'Юноши') + ', ' + distance_full_name])
                ws.merge_cells('A' + str(ws.max_row) + ':G' + str(ws.max_row))
                ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                ws.append([''])
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
                ws.append([''])
                ws.append(['', '', 'Фамилия Имя', 'г.р.', 'Команда', 'Заявка'])
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8)
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical='center')
                    if col == 4 or col == 6:
                        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                for number, swimmer in enumerate(data, start=1):
                    swimmer_ = ['', number]
                    for n, cell in enumerate(swimmer):
                        if n != 4 and n != 5 and cell != gender:
                            swimmer_.append(cell)
                        elif n == 4:
                            time = "%.2f" % cell
                            minute = str(int(cell / 60))
                            sec = f'{int(cell % 60):02}'
                            ms = str(time)[-2:]
                            swimmer_.append(return_time(cell))
                        elif cell == 'EXH':
                            swimmer_[2] += ' (EXH)'

                    ws.append(swimmer_)
                    # print(swimmer_)
                    ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                    ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                # print(('Девушки' if gender == 'ж' else 'Юноши') + ', ' + distance_full_name)

    ws_centering(ws)
    ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
    ws.oddFooter.right.size = 8
    file_name = 'Заявочный_протокол_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    close_connect_db(conn)
    wb.save(file_name)
    caption = '*Заявочный протокол ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    return file_name, caption


# def create_list(table_name):
#     file = 'список.xlsx'
#     wb = openpyxl.load_workbook(file)
#     ws = wb.active
#     conn = connect_to_db()
#     categories, sessions = tournament_year_categories(conn, table_name)
#     year_1 = 0
#     year_2 = 0
#     name_cat = ''
#     data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
#
#     ws['A3'] = data_tournament['name']
#     ws['A6'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool']
#     ws['E6'] = data_tournament['date'].strftime('%d.%m.%Y')
#     if sessions == 2:
#         ws, ws2 = create_2_sheet(wb, ws, categories, data_tournament)
#         ws.merge_cells('A2:E2')
#         ws2.merge_cells('A2:E2')
#     distances = []
#     for key, value in data_tournament.items():
#         if key.startswith("distance") and value is not None:
#             distances.append(value)
#     for session in range(1, sessions + 1):
#         if session == 2:
#             ws = ws2
#         for i, distance in enumerate(distances, start=1):
#             for cat_ in categories[session]:
#                 year_1, year_2, name_cat = years_in_categories(cat_)
#                 for gender in ['ж', 'м']:
#                     query = f"SELECT GROUP_CONCAT(second_name, ' ', first_name) AS name, gender, year, " \
#                             f"command, distance_{i}_time " \
#                             f"FROM {table_name} WHERE (year BETWEEN {year_1} AND {year_2}) AND gender = '{gender}' " \
#                             f"AND distance_{i}_time > 0 GROUP BY second_name, first_name, year, command, " \
#                             f"distance_{i}_time ORDER BY distance_{i}_time"
#                     data = select_db(conn, query, True)
#                     if data:
#                         ws.append([''])
#                         ws.append(['', distance + ',' + (' девушки' if gender == 'ж' else ' юноши') + ', ' +
#                                    name_cat + 'г.р.'])
#                         ws['B' + str(ws.max_row)].font = Font(name='Arial', size=11, bold=True)
#                         merging_cells = 'B' + str(ws.max_row) + ':D' + str(ws.max_row)
#                         ws['B' + str(ws.max_row)].alignment = Alignment(horizontal='left')
#                         ws.merge_cells(merging_cells)
#                         ws.append(['№', 'Фамилия Имя', 'г.р.', 'Команда', 'Предварительный'])
#                         ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
#                         ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='center')
#                         thins = Side(border_style="thin", )
#                         for col in range(1, 6):
#                             ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8)
#                         for number, swimmer in enumerate(data, start=1):
#                             swimmer_ = [number]
#                             for n, cell in enumerate(swimmer):
#                                 if cell != gender and n != 4:
#                                     swimmer_.append(cell)
#                                 elif n == 4:
#                                     time = "%.2f" % cell
#                                     minute = str(int(cell / 60))
#                                     sec = f'{int(cell % 60):02}'
#                                     ms = str(time)[-2:]
#                                     swimmer_.append(return_time(cell))
#                             ws.append(swimmer_)
#                             ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
#                             ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='center')
#                             ws['E' + str(ws.max_row)].alignment = Alignment(horizontal='center')
#                             for col in range(1, 6):
#                                 ws.cell(row=ws.max_row, column=col).border = Border(top=thins, bottom=thins, left=thins,
#                                                                                     right=thins)
#         ws_centering(ws)
#     close_connect_db(conn)
#     file_name = 'Список_участников_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
#
#     wb.save(file_name)
#     caption = '*Список участников ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
#               data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
#               data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
#     return file_name, caption


def create_points(table_name, file, distances):
    now = datetime.now()
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    conn = connect_to_db()
    categories, sessions = tournament_year_categories(conn, table_name)
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., ' + data_tournament['date'].strftime('%d.%m.%Y')

    distances_names = []
    distances_values = []
    for key, value in data_tournament.items():
        if value in distances:
            distances_names.append(key)
            distances_values.append(value)
    query_var = ''
    query_var_2 = ''
    for distance in distances_names:
        query_var = query_var + distance + '_fina, '
        query_var_2 = query_var_2 + distance + '_fina, '
    query_var = query_var + 'SUM('
    query_var_2 = query_var_2[:-2]
    for distance in distances_names:
        query_var = query_var + 'IF(' + distance + '_fina IS NOT NULL, ' + distance + '_fina, 0) + '
    query_var = query_var[:-3] + ')'
    for session_ in categories.keys():
        for cat_ in categories[session_]:
            year_1, year_2, name_cat = years_in_categories(cat_)
            for gender in ['ж', 'м']:
                empty_cell = '',
                ws.append(empty_cell)
                ws['B' + str(ws.max_row + 1)] = name_cat + ' ' + ('девушки' if gender == 'ж' else 'юноши')
                ws['B' + str(ws.max_row)].font = Font(name='Arial', size=10, bold=True)
                columns = ['Место', 'Участник', 'г.р.', 'Команда']
                for dist_ in distances_values:
                    columns.append(dist_)
                columns.append('Итого')
                ws.append(columns)
                for col in range(1, len(columns) + 1):
                    ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8, bold=True)
                ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center')
                for col in range(5, len(columns) + 1):
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center')
                query = f"SELECT GROUP_CONCAT(second_name, ' ', first_name) AS name, year, command, {query_var} " \
                        f"AS Total FROM {table_name} WHERE gender = '{gender}' AND (year BETWEEN {year_1} AND {year_2}) " \
                        f"GROUP BY second_name, first_name, year, command, {query_var_2} ORDER BY Total DESC"
                data_points = select_db(conn, query, True)

                for number, swimmer_ in enumerate(data_points, start=1):
                    swimmer = []
                    swimmer.append(number)
                    for column in swimmer_:
                        if column == None:
                            swimmer.append(0)
                        else:
                            swimmer.append(column)

                    ws.append(swimmer)
                    thins = Side(border_style="thin", )
                    for col in range(1, len(swimmer) + 1):
                        ws.cell(row=ws.max_row, column=col).border = Border(top=thins, bottom=thins, left=thins,
                                                                            right=thins)
                    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center')
                    ws.cell(row=ws.max_row, column=3).alignment = Alignment(horizontal='center')
                    for col in range(5, len(columns) + 1):
                        ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center')
    close_connect_db(conn)
    file_name = 'Многоборье_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    caption = '*Многоборье ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    ws_centering(ws)
    ws.oddHeader.center.text = data_tournament['name']
    ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
    ws.oddHeader.left.text = "Многоборье по очкам FINA 2022"
    ws.oddHeader.center.size = 8
    ws.oddHeader.right.size = 8
    ws.oddHeader.left.size = 8
    ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
    ws.oddFooter.right.size = 8
    wb.save(file_name)
    return file_name, caption


def statistic_years(table_name):
    conn = connect_to_db()
    wb = openpyxl.load_workbook("участники.xlsx")
    ws = wb.active
    thins = Side(border_style="thin", )
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., ' + data_tournament['date'].strftime('%d.%m.%Y')

    ws.append([''])
    ws.append(['', 'Команда', '', '', '', 'Юноши', 'Девушки', 'Всего'])
    ws['H' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['G' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=9)
        ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
    query = f"SELECT DISTINCT command FROM {table_name} ORDER BY command"
    data_commands = select_db(conn, query, True)
    all_commands = []
    for y in data_commands:
        all_commands.append(y[0])
    boys = 0
    girls = 0
    for com in all_commands:
        data = ['']
        data.append(com)
        data.append('')
        data.append('')
        data.append('')
        for g in ['м', 'ж']:
            query = f"SELECT count(gender) AS count FROM {table_name} WHERE command = '{com}' AND gender = '{g}'"
            count = select_db(conn, query, False)
            data.append(count['count'])
            if g == 'м':
                boys += count['count']
            else:
                girls += count['count']
        data.append(data[5] + data[6])
        ws.append(data)
        ws['H' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['G' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws.append([''])
    for col in range(1, 9):
        ws.cell(row=ws.max_row, column=col).border = Border(top=thins)
    ws.append(['', 'Всего', '', '', '', boys, girls, boys + girls])
    ws['H' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['G' + str(ws.max_row)].alignment = Alignment(horizontal='center')

    ws.append([''])
    ws.append(['', 'Год рождения', 'Юноши', 'Девушки', 'Всего'])
    ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    ws['E' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    for col in range(1, 6):
        ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=9)
        ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
    query = f"SELECT DISTINCT year FROM {table_name} ORDER BY year"
    data_years = select_db(conn, query, True)
    all_years = []
    for y in data_years:
        all_years.append(y[0])
    # print(all_years)

    for y in all_years:
        data = ['']
        data.append(y)
        for g in ['м', 'ж']:
            query = f"SELECT count(gender) AS count FROM {table_name} WHERE year = {y} AND gender = '{g}'"
            count = select_db(conn, query, False)
            data.append(count['count'])
        data.append(data[2] + data[3])
        # print(data)
        ws.append(data)
        ws['B' + str(ws.max_row)].alignment = Alignment(horizontal='left')
        ws['C' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
        ws['E' + str(ws.max_row)].alignment = Alignment(horizontal='center')

    count_distance = 0
    for key, value in data_tournament.items():
        if key.startswith("distance") and value is not None:
            count_distance += 1
    query = f"SELECT null, GROUP_CONCAT(second_name, ' ', first_name) AS name, null, year, " \
            f"command FROM {table_name} WHERE"
    for r in range(1, count_distance + 1):
        query += f" distance_{str(r)}_dsq = 'EXH' or"
    query = query[:-2] + "GROUP BY second_name, first_name, year, command"
    # print(query)
    data_exh = select_db(conn, query, True)
    # print(data_exh)
    if data_exh:
        ws.append([''])
        ws.append(['', 'Лично (EXH)'])
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).border = Border(bottom=thins)
    for swimmer in data_exh:
        ws.append(swimmer)
        ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
    # print(data_exh)

    close_connect_db(conn)
    file_name = 'Участники_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    caption = '*Участники ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    ws_centering(ws)
    wb.save(file_name)
    return file_name, caption


def statistic_by_cat(table_name):
    global __last_cell_of_page__
    file = 'заявочный.xlsx'
    now = datetime.now()
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    ws['A1'] = "Список участников по категориям"
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., '
    distances = []
    thins = Side(border_style="thin", )
    for key, value in data_tournament.items():
        if key.startswith("distance") and value is not None:
            distances.append(value)
    categories, sessions = tournament_year_categories(conn, table_name)
    # print(categories)
    for i, distance in enumerate(distances, start=1):
        distance_full_name = full_name_distance(distance)
        for session_ in categories.keys():
            for cat_ in categories[session_]:

                year_1, year_2, name_cat = years_in_categories(cat_)
                for gender in ['ж', 'м']:
                    query = f"SELECT GROUP_CONCAT(second_name, ' ', first_name) AS name, gender, year, " \
                            f"command, distance_{i}_time, distance_{i}_dsq " \
                            f"FROM {table_name} WHERE (year BETWEEN {year_1} AND {year_2}) AND gender = '{gender}' " \
                            f"AND distance_{i}_time > 0 GROUP BY second_name, first_name, year, command, " \
                            f"distance_{i}_time, distance_{i}_dsq ORDER BY name"
                    # print(query)
                    data = select_db(conn, query, True)
                    # print(data)
                    if data:
                        max_row = ws.max_row
                        list_of_cells_numbers = []
                        for h in range(1, len(data) + 6):
                            list_of_cells_numbers.append(h + max_row)
                        if __last_cell_of_page__ in list_of_cells_numbers:
                            page_break = Break(id=max_row)
                            __last_cell_of_page__ = max_row + 61
                            ws.row_breaks.append(page_break)
                        elif list_of_cells_numbers[0] > __last_cell_of_page__:
                            __last_cell_of_page__ = max_row + 61
                        ws.append([''])
                        ws.append(
                            [distance_full_name + (', девушки' if gender == 'ж' else ', юноши') + ', ' + name_cat])
                        ws.merge_cells('A' + str(ws.max_row) + ':G' + str(ws.max_row))
                        ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                        for col in range(1, 8):
                            ws.cell(row=ws.max_row, column=col).border = Border(top=thins, bottom=thins)
                        ws.append([''])
                        ws.append(['', '', 'Фамилия Имя', 'г.р.', 'Команда', 'Заявка'])
                        for col in range(1, 8):
                            ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8)
                            ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical='center')
                            if col == 4 or col == 6:
                                ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                        for number, swimmer in enumerate(data, start=1):
                            swimmer_ = ['', number]
                            for n, cell in enumerate(swimmer):
                                if n != 4 and n != 5 and cell != gender:
                                    swimmer_.append(cell)
                                elif n == 4:
                                    time = "%.2f" % cell
                                    minute = str(int(cell / 60))
                                    sec = f'{int(cell % 60):02}'
                                    ms = str(time)[-2:]
                                    swimmer_.append(return_time(cell))
                                elif cell == 'EXH':
                                    swimmer_[2] += ' (EXH)'

                            ws.append(swimmer_)
                            # print(swimmer_)
                            ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                            ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                    # print(('Девушки' if gender == 'ж' else 'Юноши') + ', ' + distance_full_name)
    __last_cell_of_page__ = 61
    ws_centering(ws)
    ws.oddHeader.center.text = data_tournament['name']
    ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
    ws.oddHeader.left.text = "Список участников"
    ws.oddHeader.center.size = 8
    ws.oddHeader.right.size = 8
    ws.oddHeader.left.size = 8
    ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
    ws.oddFooter.right.size = 8
    file_name = 'Список_участников_по_категориям_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    close_connect_db(conn)
    wb.save(file_name)
    caption = '*Список участников по категориям ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    return file_name, caption


def statistic_by_com(table_name):
    global __last_cell_of_page__
    file = 'заявочный.xlsx'
    now = datetime.now()
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    conn = connect_to_db()
    data_tournament = select_db(conn, None, False, 'tournaments', '*', table_name="'" + table_name + "'")
    ws['A1'] = "Список участников по категориям"
    ws['A2'] = data_tournament['name']
    ws['A3'] = data_tournament['place'] + ', ' + data_tournament['swimming_pool'] + ', ' + \
               str(data_tournament['tracks']) + ' дор., '
    distances = []
    thins = Side(border_style="thin", )
    for key, value in data_tournament.items():
        if key.startswith("distance") and value is not None:
            distances.append(value)
    categories, sessions = tournament_year_categories(conn, table_name)
    # print(categories)
    for i, distance in enumerate(distances, start=1):
        distance_full_name = full_name_distance(distance)
        for session_ in categories.keys():
            for cat_ in categories[session_]:

                year_1, year_2, name_cat = years_in_categories(cat_)
                for gender in ['ж', 'м']:
                    query = f"SELECT GROUP_CONCAT(second_name, ' ', first_name) AS name, gender, year, " \
                            f"command, distance_{i}_time, distance_{i}_dsq " \
                            f"FROM {table_name} WHERE (year BETWEEN {year_1} AND {year_2}) AND gender = '{gender}' " \
                            f"AND distance_{i}_time > 0 GROUP BY second_name, first_name, year, command, " \
                            f"distance_{i}_time, distance_{i}_dsq ORDER BY name"
                    # print(query)
                    data = select_db(conn, query, True)
                    # print(data)
                    if data:
                        max_row = ws.max_row
                        list_of_cells_numbers = []
                        for h in range(1, len(data) + 4):
                            list_of_cells_numbers.append(h + max_row)
                        if __last_cell_of_page__ in list_of_cells_numbers:
                            page_break = Break(id=max_row)
                            __last_cell_of_page__ = max_row + 61
                            ws.row_breaks.append(page_break)
                        elif list_of_cells_numbers[0] > __last_cell_of_page__:
                            __last_cell_of_page__ = max_row + 61
                        ws.append([''])
                        ws.append(
                            [distance_full_name + (', девушки' if gender == 'ж' else ', юноши') + ', ' + name_cat])
                        ws.merge_cells('A' + str(ws.max_row) + ':G' + str(ws.max_row))
                        ws['A' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                        for col in range(1, 8):
                            ws.cell(row=ws.max_row, column=col).border = Border(top=thins, bottom=thins)
                        ws.append([''])
                        ws.append(['', '', 'Фамилия Имя', 'г.р.', 'Команда', 'Заявка'])
                        for col in range(1, 8):
                            ws.cell(row=ws.max_row, column=col).font = Font(name='Arial', size=8)
                            ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical='center')
                            if col == 4 or col == 6:
                                ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                        for number, swimmer in enumerate(data, start=1):
                            swimmer_ = ['', number]
                            for n, cell in enumerate(swimmer):
                                if n != 4 and n != 5 and cell != gender:
                                    swimmer_.append(cell)
                                elif n == 4:
                                    time = "%.2f" % cell
                                    minute = str(int(cell / 60))
                                    sec = f'{int(cell % 60):02}'
                                    ms = str(time)[-2:]
                                    swimmer_.append(return_time(cell))
                                elif cell == 'EXH':
                                    swimmer_[2] += ' (EXH)'

                            ws.append(swimmer_)
                            # print(swimmer_)
                            ws['D' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                            ws['F' + str(ws.max_row)].alignment = Alignment(horizontal='center')
                    # print(('Девушки' if gender == 'ж' else 'Юноши') + ', ' + distance_full_name)
    __last_cell_of_page__ = 61
    ws_centering(ws)
    ws.oddHeader.center.text = data_tournament['name']
    ws.oddHeader.right.text = data_tournament['date'].strftime('%d.%m.%Y')
    ws.oddHeader.left.text = "Список участников"
    ws.oddHeader.center.size = 8
    ws.oddHeader.right.size = 8
    ws.oddHeader.left.size = 8
    ws.oddFooter.right.text = now.strftime(' %d.%m.%Y %H:%M') + "  Страница &P из &N"
    ws.oddFooter.right.size = 8
    file_name = 'Список_участников_по_категориям_' + data_tournament['date'].strftime('%d_%m_%Y') + '.xlsx'
    close_connect_db(conn)
    wb.save(file_name)
    caption = '*Список участников по категориям ' + data_tournament['date'].strftime('%d.%m.%Y') + '*\n' + \
              data_tournament['name'] + '\n' + data_tournament['place'] + ', ' + \
              data_tournament['swimming_pool'] + ', ' + str(data_tournament['tracks']) + ' дор.\n'
    return file_name, caption


def years_in_categories(cat_):
    year_1 = 0
    year_2 = 0
    name_cat = ''
    if len(cat_) > 4:
        if cat_[-1] == '+':
            year_1 = 1950
            year_2 = cat_[:-1]
            name_cat = str(year_2) + ' и старше'
        elif cat_[-1] == '-':
            year_1 = cat_[:-1]
            year_2 = 2025
            name_cat = str(year_1) + ' и младше'
        elif cat_[4] == '-':
            year_1 = cat_[:4]
            year_2 = cat_[5:]
            name_cat = str(year_1) + '-' + str(year_2)
    elif len(cat_) == 4:
        year_1 = cat_
        year_2 = cat_
        name_cat = str(year_1)
    return year_1, year_2, name_cat
