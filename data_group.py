import openpyxl
import datetime as dt
from db import select_db, connect_to_db, update_db, close_connect_db
from sport_categories import fina_points


def write_to_group_open_excel(table_name):
    wb = openpyxl.load_workbook(table_name + '.xlsx')
    ws = wb.active
    message_text = []
    distance_row_number = []
    bool_ = False
    for row in range(1, ws.max_row + 1):
        if ws['A' + str(row)].value is not None and ws['A' + str(row)].value.startswith("Дистанция"):
            distance_row_number.append(row)
    for n, num in enumerate(distance_row_number):
        if n == len(distance_row_number) - 1:
            dist_row_ = ws.max_row
        else:
            dist_row_ = distance_row_number[n + 1]
        for row in range(num, dist_row_):
            if ws['B' + str(row)].value is not None and ws['C' + str(row)].value is not None and ws[
                'G' + str(row)].value is None:
                distance = ws['A' + str(num)].value + '\n' + ws['D' + str(num)].value
                message_text.append(distance)
                bool_ = True
                break

        for row in range(num, dist_row_):
            if ws['B' + str(row)].value is not None and ws['C' + str(row)].value is not None and ws[
                'G' + str(row)].value is None:
                for ss in range(row, 1, -1):
                    if ws['C' + str(ss)].value is not None and ws['C' + str(ss)].value.startswith('Заплыв'):
                        swim = ws['C' + str(ss)].value
                        message_text.append(swim)
                        s = ss
                        break
                break
        if message_text:
            for row in range(num, dist_row_):
                if ws['C' + str(row)].value == message_text[1]:
                    for s in range(row + 1, row + 10):
                        if ws['B' + str(s)].value is not None and ws['C' + str(s)].value is not None:
                            message_text.append(
                                str(ws['B' + str(s)].value) + ' дорожка:\n' + ws['C' + str(s)].value + ', ' + str(
                                    ws['D' + str(s)].value) + '/' + str(ws['E' + str(s)].value))
                        if ws['C' + str(s)].value is not None and ws['C' + str(s)].value.startswith('Заплыв'):
                            break

        if bool_:
            break
    conn = connect_to_db()
    query = f"INSERT IGNORE INTO temp_protocol(table_name, sending, receive) " \
            f"VALUES ('{table_name}', {len(message_text) - 2}, 0)"
    cursor = conn.cursor(buffered=True)
    cursor.execute(query)
    conn.commit()
    close_connect_db(conn)
    return message_text


def group_write_to_excel_and_db(group_id, text, result_str):
    text = text.split('/')
    name = text[0].split(', ')[0]
    second_name = name[:name.index(' ')]
    first_name = name[name.index(' ') + 1:]
    command = text[1]
    distances = []
    number_distance = 0
    conn = connect_to_db()
    query = f"SELECT * FROM tournaments WHERE id_telegram=" \
            f"(SELECT id_telegram FROM users WHERE group_id = {group_id}) and date = '{dt.date.today()}'"
    data_tournament = select_db(conn, query, False)
    print(data_tournament)
    try:
        table_name = data_tournament['table_name']
    except:
        return 'На сегодняшний день турниров не найдено.'
    wb = openpyxl.load_workbook(table_name + '.xlsx')
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        if ws['C' + str(row)].value == name and ws['E' + str(row)].value == command:
            ws['G' + str(row)] = result_str
            wb.save(table_name + '.xlsx')
            break

    for r in range(row, 1, -1):
        if ws['A' + str(r)].value is not None and ws['A' + str(r)].value.startswith('Дистанция'):
            distance = ws['D' + str(r)].value
            if "девушки" in distance:
                gender = 'ж'
            elif "юноши" in distance:
                gender = 'м'
            distance = distance[:distance.index(',')]
            break
    for key, value in data_tournament.items():
        if key.startswith("distance"):
            if value is not None:
                distances.append(value)
    for i, value in enumerate(distances, start=1):
        if value == distance:
            number_distance = i
            break
    if result_str and ',' in result_str:
        minute = int(result_str[:result_str.index('.')])
        sec = int(result_str[result_str.index('.') + 1:result_str.index(',')])
        ms = int(result_str[result_str.index(',') + 1:])
        result = minute * 60 + sec + ms / 100
    else:
        result = 9999.0

    if "(EXH)" in first_name:
        first_name = first_name[:first_name.index(" (EXH)")]
        set_query = f"distance_{str(number_distance)}_dsq='EXH', "
    else:
        set_query = f"distance_{str(number_distance)}_dsq=NULL, "
    set_query += f"distance_{str(number_distance)}_result={result}, " \
                f"distance_{str(number_distance)}_fina=" \
                f"{fina_points(gender, data_tournament['swimming_pool'], distance, result)}"

    query = f"UPDATE {table_name} SET {set_query} WHERE second_name='{second_name}' " \
            f"AND first_name='{first_name}' AND command='{command}'"
    print(query)
    update_db(conn, query)
    query = f"UPDATE temp_protocol SET receive = receive + 1 WHERE table_name = '{table_name}'"
    print(query)
    update_db(conn, query)
    rows = select_db(conn, None, False, 'temp_protocol', 'sending', 'receive', table_name="'" + table_name + "'")
    swim_bool = False
    if rows['sending'] == rows['receive']:
        query = f"DELETE FROM `temp_protocol` WHERE table_name = '{table_name}'"
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        swim_bool = True
    close_connect_db(conn)

    return 'Результат добавлен', swim_bool
